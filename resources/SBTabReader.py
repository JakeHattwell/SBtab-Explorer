#!/usr/bin/env python

"""SBTabReader.py - converts a csv with WormBase RNAi identifiers in the first column to a list of corresponding target genes

Requires (use pip to install):
openpyxl
"""
import os
import time
import tkinter as tk
import csv

import openpyxl

__author__ = "Jake Hattwell"
__copyright__ = "None"
__credits__ = ["Jake Hattwell"]
__license__ = "CCO"
__version__ = "1"
__maintainer__ = "Jake Hattwell"
__email__ = "j.hattwell@uq.edu.au"
__status__ = "Live"


class modelSystem():
    """Class for use with SBTabInterface.py
    """
    def __init__(self,master=None):
        self.tables = {}
        self.master=master
        self.size = {}
    
    def loadTable(self,name,location,filetype):
        self.tables[name] = dataset(location,mode=filetype)
        self.size[name] = self.tables[name].rows-2

    def load_folder(self,name,filetype):
        success = False
        self.master.print_out("------------------------")
        if os.path.isdir(name) == False:
            self.master.print_out("This folder does not exist")

        else:
            self.master.print_out("Folder loaded")
            paths = []
            for f in os.listdir(name):
                if filetype == "xlsx":
                    if "SBtab.xlsx" in f:
                        if name[1] != ":":
                            path = os.getcwd()+"\\"+name+"\\"+f
                        else:
                            path = name+"\\"+f
                            filename = f.replace("-SBtab.xlsx","")
                            filename = f.replace("-SBtab.tsv","")
                            paths.append([filename,path])
                elif filetype == "tsv":
                    if "SBtab.tsv" in f:
                        if name[1] != ":":
                            path = os.getcwd()+"\\"+name+"\\"+f
                        else:
                            path = name+"\\"+f
                            filename = f.replace("-SBtab.xlsx","")
                            filename = f.replace("-SBtab.tsv","")
                            paths.append([filename,path])

            if paths == []:
                self.master.print_out(" ".join(["There were no SBtab files found in",name]))
            else:
                self.master.print_out("SBtab files found! Loading now!")
                self.count=1
                
                for hit in paths:
                    
                    self.master.print_out(" ".join(["Loading file:",hit[0]]))
                    self.loadTable(hit[0],hit[1],filetype)
                    self.count+=1
                    self.master.footer.config(width=self.master.master.winfo_width()*self.count/len(paths),bg="green2")
                    self.master.placeholder.config(width=self.master.master.winfo_width()*(len(paths)-self.count)/len(paths))
                    time.sleep(0.05)

                self.master.print_out(" ".join([str(len(paths)),"files loaded into the model"]))
                success = True    
            
        return success
                    

    def searchModel(self,term,dataset="All",mute=False):
        results = {}
        count = 0
        for table,contents in self.tables.items():
            for ID,entry in contents.data.items():
                for key,val in entry.items():
                    try:
                        if term.lower() in str(val).lower():
                            row = list(contents.data).index(ID) + 3
                            results["-".join([table,ID])] = [table,ID,key,str(val),str(row),entry]
                            count += 1
                    except:
                        self.master.print_out("Error searching for term")
        if count != 0 and mute==False:
            self.master.print_out("------------------------")
            
            self.master.print_out(" ".join([str(len(results)),"hits found!"]))
        elif mute == False:
            self.master.print_out(" ".join(["Search term",term,"returned 0 results"]))
        return results
    
    def prettyPrint(self,accession):
        data = self.tables[accession[0]].data[accession[1]]
        output = ""
        output += " ".join([accession[0]+":",accession[1],"\n"])
        lineCount = 0
        for iden,entry in data.items():
            if iden != None and lineCount < 5 and lineCount > 0:
                output += " ".join([str(iden)+":",str(entry),"\n"])
            lineCount += 1
        output = output[:-2]
        return output
    

class dataset:
    """Importable class for loading SBTab files\nConverts SBTab as nested dictionary.\n

    instance.data = Dictionary of entries in SBTab\n
    Each entry is a dictionary of the data associated with that entry, with column headers as keys.
        
        Arguments:
            xlsx {str} -- Path to SBTab file of interest.
        
        Keyword Arguments:
            headerRow {int} -- Excel row of the header information, (default: {2})
            mode {str} -- version of dataset to load
        """

    def __init__(self,filename,headerRow=2,mode="xslx"):
        """Loads the SBTab file"""
        self.name = filename
        if mode=="xlsx":
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            self.cols = sheet.max_column
            self.rows = sheet.max_row
            self.sbString = sheet.cell(row=1,column=1).value
            try:
                self.headers = [sheet.cell(row=2,column = i).value for i in range(1,self.cols+1) if sheet.cell(row=headerRow,column = i)!= None]
                self.data = {str(sheet.cell(row=i,column = 1).value):{self.headers[j-1]:sheet.cell(row=i,column=j).value for j in range(1,self.cols+1)}for i in range(headerRow+1,self.rows+1)}
                self.freeze_panes = sheet.freeze_panes
            except:
                print(self.name)
                print("XLSX import failed. Aborting...")
                exit()
        elif mode=="tsv":
            with open(filename,encoding="utf-8") as tsvfile:
                tsv = csv.reader(tsvfile,delimiter="\t")
                entries = []
                for row in tsv:
                    if tsv.line_num == 1: #row 1 - SBtab DocString
                        self.sbString = row[0]
                    elif tsv.line_num == 2: #row 2 - headers of the table
                        self.headers = row
                    else:
                        entries.append(row)
            # define size of data
            self.cols = len(self.headers)
            self.rows = len(entries)+2
            # create the nested dict object
            try:
                self.data = {entry[0]:{self.headers[i]:(entry[i] if len(entry) >= len(self.headers) else '') for i in range(1,len(self.headers))} for entry in entries}
                while '' in self.data:
                    self.data.pop('')
            except:
                print(self.name)
                print("tsv import failed. Aborting...")
                exit()
            #remove blank entries


                

    def saveToExcel(self,name):
        newWb = openpyxl.Workbook()
        newWs = newWb.active
        newWs['A1'] = self.sbString
        for i in range(len(self.headers)):
            newWs.cell(row=2,column=i+1).value = self.headers[i]
        row = 3
        for key,val in self.data.items():
            del key
            col = 1
            for i in self.headers:
                newWs.cell(row=row,column=col).value = val[i]
                col += 1
            row += 1
        if hasattr(self,"freeze_panes"):
            newWs.freeze_panes = self.freeze_panes
        newWb.save(name+'-SBtab.xlsx')
